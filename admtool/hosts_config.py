'''  coding: utf-8
zabbix a hosts configuration statistic excel file with : 
- host name / description / visible ...
- specific host user macros
- specific host tag value (zabbix version >= 4.2)
- specific host inventory attributes


Author : Francois Martin-Lefevre : fml@axynergie.com


'''

import sys
import os
import argparse
import filecmp
import re
import tempfile
import itertools as it

from pyzabbix.api import ZabbixAPI, ZabbixAPIException

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle, PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


from zbxtool_functions import *

 
def parse_arg():
    parser = argparse.ArgumentParser(add_help=True, description='zbxtools :: statistics of hosts configuration loading ')
    parser.add_argument("--config", type=str,
                        help="config file path and name (default: config.ini)", 
                        default="config.ini")
    parser.add_argument("--verbose", type=bool,
                        help="mode verbose (default: false)",
                        default=False)
    parser.add_argument("--zbxenv", type=str,
                        help="config file zabbix environment section")
    parser.add_argument("--xlfile", type=str,
                        help="name of excel output file")
    parser.add_argument("--limit", type=int,
                        help="for test purpose, limits the number of host scanned")


    args=parser.parse_args()

    return args


def zbx_fmt_host(zapi,host,convert, proxy_name):
    ''' return a formated host definition wich convert 
        - proxy_id with corresponding proxy name
        - suppress 'hostid' if not in selection value
        - all values indicated in convert table
    '''
    result = {}
    tmp_result = dict(host)

    # loop on specific items and replace corresponding value
    for key1 in host:
        if key1 == "proxy_hostid": 
            try:
                tmp_result[key1] = proxy_name[host[key1]]
            except KeyError:
                continue
        elif key1 == "hostid":
            if key1 in sel["hosts"]:
                continue
            else:
                del tmp_result[key1]
        elif key1 == "inventory":
            try:
                tmp_result[key1]["os_full"] = tmp_result[key1]["os_full"].replace(tmp_result[key1]["name"], "") 
            except KeyError:
                continue
        else:
            try:
                tmp_result[key1] = convert[key1][host[key1]]
            except KeyError:
                continue

    result.update(tmp_result)

    return result

def zbx_proxy_name(zapi):
    ''' query the declared proxy and construct a dictionnary with
        key = proxyid  / value = host
    '''
    result = {}
    proxy_qry = zapi.do_request('proxy.get',
                    {"output": ["proxyids", "host"],
                    })['result']
    for proxy in proxy_qry:
        result.update({proxy["proxyid"]:proxy["host"]})
    
    return result

def zbx_host_config(zapi, hosts):
    ''' return a list of hosts :
        - attributes : as values of dict
        - macros, groups, templates, inventory, tags(*) : as list of dict
        
        * tags only with zabbix > 4.2.0
    '''

    result = []
    host = []
    global sel
    global desc

    convert = zbx_table("init")

    # query the proxy dictionnary
    proxy_name = zbx_proxy_name(zapi)

    idx_host = 0


    for ho in hosts:
        idx_host += 1    

        if v_zabbix >=420:
            host = zapi.do_request('host.get',
                    {"output": sel["hosts"],
                    "hostids": ho["hostid"],
                    "selectMacros": sel["macros"],
                    "selectGroups": sel["groups"],
                    "selectParentTemplates": sel["templates"],
                    "selectInventory": sel["inventory"],
                    # "selectItems" : sel["items"],
                    "selectTags": sel["tags"],
                    })['result'][0]
        else:
            host = zapi.do_request('host.get',
                    {"output": sel["hosts"],
                    "hostids": ho["hostid"],
                    "selectMacros": sel["macros"],
                    "selectGroups": sel["groups"],
                    "selectParentTemplates": sel["templates"],
                    "selectInventory": sel["inventory"],
                    # "selectItems" : sel["items"],
                    })['result'][0]
        # print(host[0]["host"])


        logging.info("host (" + str(idx_host) + "/"  + str(len(hosts)) + ") : " + host["host"])

        fmt_host = zbx_fmt_host(zapi, host, convert, proxy_name)

        result.append(fmt_host)

        if (limit_host != 0 and idx_host >= limit_host):
            logging.info("ended program due to host scan number limitation (--limit parameter)")
            break

    return result

def extract_list(hosts_tbl, it2, it3):
    ''' extract list of value in the host dictionnary
            struct of hosts_tbl = {it1:value, it2:[{it3:value}, ...], ...}
        it2 : dictionnary name of the list
        it3 : item name in the dictionnary list
        the return list is a list of it3 values
        if it2 or it3 don't exist -> return a blank list
    '''
    result = []

    try:
        for ho in hosts_tbl:
            for idx in ho[it2]:
                if idx[it3] not in result:
                    result.append(idx[it3])
    except KeyError:
        pass
    
    return result

def extract_noempty_list(hosts_tbl):
    ''' extract key of no empty values in inventory dictionnary 

    '''
    result = []
    for ho in hosts_tbl:
        for key in ho["inventory"]:
            if key != "hostid" and key not in result and ho["inventory"][key] != "":
                result.append(key)

    return result


def extract_lists(hosts_tbl):
    ''' extract the lists of macros, tags, templates, groups, inventory

    '''
    result = {}

    macros_list = extract_list(hosts_tbl,"macros","macro")
    result.update({"macros": macros_list})

    tags_list = extract_list(hosts_tbl, "tags", "tag")
    result.update({"tags": tags_list})

    templates_list = extract_list(hosts_tbl, "parentTemplates", "host")
    result.update({"parentTemplates": templates_list})
    
    groups_list = extract_list(hosts_tbl, "groups", "name")
    result.update({"groups": groups_list})

    inventory_list = extract_noempty_list(hosts_tbl)
    result.update({"inventory": inventory_list})


    return result

def ws_host_title(host_tbl, host_lst):
    ''' format host title for excel output
        columns title are prefixed with 
            ho: host data
            ma: macros
            ta: tags
            in: inventory
        for host data the title is replaced with desc attributes
    '''

    result =[]
    global desc

    title = dict(zip(sel["hosts"], desc["hosts"]))

    for key in host_tbl[0]:
        if key not in ["groups", "parentTemplates", "macros", "tags", "inventory"]:
            try: 
                result.append("ho_" + title[key])
            except KeyError:
                pass
        elif key == "macros":
            for idx in host_lst["macros"]:
                result.append("ma_" + idx)
        elif key == "tags":
            for idx in host_lst["tags"]:
                result.append("ta_" + idx)
        elif key == "inventory":
            for key in host_lst["inventory"]:
                result.append("in_" + key)

    return result

def ws_host_data(host_tbl, host_lst):
    ''' append zabbix data to worksheet data structure
            [[row1],[row2], ...]
        where row is a list of 'host_tpl' value in the correct 'host_lst' order
    '''

    result =[]
    for ho in host_tbl:
        tmp_result=[]
        for key in ho:
            if key not in ["groups", "parentTemplates", "macros", "tags", "inventory", "items"]:
                tmp_result.append(ho[key])
            elif key == "macros":
                for id1 in host_lst["macros"]:
                    tmp_val = ""
                    for id2 in ho[key]:
                        if id1 == id2["macro"]:
                            tmp_val = id2["value"]
                    tmp_result.append(tmp_val)
            elif key == "tags":
                for id1 in host_lst["tags"]:
                    tmp_val = ""
                    for id2 in ho[key]:
                        if id1 == id2["tag"]:
                            tmp_val = id2["value"]
                    tmp_result.append(tmp_val)
            elif key == "inventory":
                for key1 in host_lst["inventory"]:
                    try:
                        tmp_result.append(ho["inventory"][key1])
                    except KeyError:
                        tmp_result.append("")

        result.append(tmp_result)

    return result

def ws_group_data(host_data, host_lst):
    result =[]
    for ho in host_data:
        for id1 in ho["groups"]:
            result.append([ho["host"],id1["name"]])

    return result

def ws_template_data(host_data, host_lst):
    result =[]
    for ho in host_data:
        for id1 in ho["parentTemplates"]:
            result.append([ho["host"],id1["host"]])

    return result

def xl_zbx_create(wb, wsname, wstitle,wsdata):
    ''' create sheet and apply format


    '''
    ws = wb.create_sheet(wsname)
    ws.append(wstitle)
    for row in wsdata:
        ws.append(row)

    first_cell = ws.cell(row=1,column=1).coordinate
    last_cell = ws.cell(row=len(wsdata)+1,column=len(wstitle)).coordinate
    refstr = first_cell + ":" + last_cell
    # print(refstr)

    tab = Table(displayName=wsname, ref=refstr)
    style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    return ws


def xl_create_dataframe(ws):
    ''' create a pandas dataframe from a ws sheet
        - first line of ws sheet is the columns name
        - no index in the ws, values in columns are not unique

        return a pandas dataframe object
    '''
    data = ws.values
    cols = next(data)[0:]
    df = pd.DataFrame(data, columns=cols)

    return df

def zbx_stat(df, ws, val, pv_lst, startcol):
    ''' store pivot statistics in specific worksheet
        inputs :
            df : dataframe
            ws : worksheet
            val : count value for pivot
            pv_list : list of columns pivot compute
            statcol : starting column for value writing (the row is allways set to 1) 
    '''
    for pv in pv_lst:
        pivot = df.pivot_table(index=pv, values=val, aggfunc="count")
        if mode_verbose:
            logging.info("stat process -> %s statistics", pv)
        xl_stat(ws, pivot, startcol)
        startcol += 3

    return


def xl_stat(ws, pv, startcol):
    ''' format all statistics in a single sheet 
        inputs :
            - worksheet name (if not exist will be created)
            - pv : pivot dataframe  
            - startcolumn
    '''
    for id_row, r in enumerate(dataframe_to_rows(pv)):
        row = list(r)
        for id_col, val in enumerate(row):
            # print("r", id_row, "c", id_col, "val", val)
            if id_row == 0 and id_col == 1: 
                ws.cell(row=1, column=startcol+1).value = val
                ws.cell(row=1, column=startcol+1).style = "zbx title"
            elif id_row == 1 and id_col == 0:
                ws.cell(row=1, column=startcol).value = val
                ws.cell(row=1, column=startcol).style = "zbx title"
            elif id_row >=1 and id_col >= 0:
                ws.cell(row=id_row, column=startcol+id_col).value = val    
                ws.cell(row=id_row, column=startcol+id_col).style = "zbx data"    
    return


if __name__ == '__main__':

    module = __file__[0:__file__.find(".")]

    ''' process args and config parameters for module variables setting
        1/ args parameter
        2/ config parameters
        3/ if config parameter not exist then args is the default
    '''

    args = parse_arg()
    config = parse_config(args,module)

    mode_verbose = args.verbose or config.getboolean("default", "mode_verbose") or False
    init_log(config, mode_verbose)

    zbxenv = args.zbxenv or config["default"]["env"] or "Zabbix"
    limit_host = args.limit or 0

    xlfile = args.xlfile or config[module]["xlfile"] or "host.xlsx"

    zbxtool_dir = config["paths"]["zbxtool_dir"]  
    dir_data = zbxtool_dir + "/"  + config[module]["save_dir"] + "/"  + zbxenv + "/"

    logging.info("save directory is : %s", dir_data)

    if not os.path.exists(dir_data):
        os.makedirs(dir_data)

    if limit_host != 0:
        logging.info("max number of host activated %s hosts", str(limit_host))

    
    logging.info("zabbix environment = %s ", zbxenv)


    # init global variable
    global sel
    sel = {}
    global desc
    desc = {}

    # start of program logic
    zapi = zbx_connect(config,zbxenv)
    v_zabbix = int(zapi.api_version().replace(".",""))

    # collect all hostsid ordered by hostname
    hosts = zapi.do_request('host.get',{"output": ["hostid"],
                "sortfield" : "host",
                })['result']

    # loop on the hostids for collecting the selectionned data
    sel.update({"hosts": ["host", "description","name", "proxy_hostid", "flags", "maintenance_status",
                        "status", "snmp_available"]})
    desc.update({"hosts": ["host", "description", "name", "proxy", "mode ajout", "maintenance",
                        "etat", "interface snmp"]})
    sel.update({"macros": ["macro", "value"]})
    sel.update({"groups": ["name"]})
    sel.update({"templates": ["host"]})
    sel.update({"inventory": "extend"})
    sel.update({"tags": ["tag", "value"]})
    sel.update({"items": ["itemid", "name", "_key", "type"]})
    hosts_tbl = zbx_host_config(zapi, hosts)


    # extract lists
    zbx_host_list = extract_lists(hosts_tbl)
    # print(zbx_host_list)
    
    # create and process excel data 
    wb = Workbook()

    # create named style for excel statistics sheet formating
    zbx_title = NamedStyle(name="zbx title")
    zbx_title.font = Font(bold=True, color="ffffff")
    zbx_title.fill = PatternFill(fill_type="solid", start_color="538DD5", end_color="538DD5")
    bd = Side(style="thin", color="538DD5")
    zbx_title.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    wb.add_named_style(zbx_title)

    zbx_data = NamedStyle(name="zbx data")
    zbx_data.font = Font(color="000000")
    zbx_data.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    wb.add_named_style(zbx_data)


    # process host data
    host_title =  ws_host_title(hosts_tbl, zbx_host_list)
    host_data = ws_host_data(hosts_tbl, zbx_host_list)
    ws_host = xl_zbx_create(wb, "Hosts", host_title, host_data)
 
    logging.info("sheet 'Hosts' created : " + str(len(host_title)) 
            + " cols " + str(len(host_data)) + " rows")

    # process groups data
    group_title = ["host", "group"]
    group_data = ws_group_data(hosts_tbl, zbx_host_list)
    ws_groups = xl_zbx_create(wb, "Groups", group_title, group_data)
 
    logging.info("sheet 'Groups' created : " + str(len(group_title)) 
            + " cols " + str(len(group_data)) + " rows")


    # process templates data
    template_title = ["host", "template"]
    template_data = ws_template_data(hosts_tbl, zbx_host_list)
    ws_templates = xl_zbx_create(wb, "Templates", template_title, template_data)

    logging.info("sheet 'Templates' created : " + str(len(template_title)) 
            + " cols " + str(len(template_data)) + " rows")

    wb.save(dir_data + xlfile)

    # process statistics data -> create pandas dataframe and proceed with pivot table
    ws_stat = wb.create_sheet("STATS")

    logging.info("process 'groups' statistics data") 
    df_groups = xl_create_dataframe(ws_groups)
    groups_pivot = ["group"]
    groups_value = "host"
    last_col = 1
    zbx_stat(df_groups, ws_stat, groups_value, groups_pivot, last_col)
    last_col += len(groups_pivot) * 3

    logging.info("process 'templates' statistics data") 
    df_templates = xl_create_dataframe(ws_templates)
    templates_pivot = ["template"]
    templates_value = "host"
    zbx_stat(df_templates, ws_stat, templates_value, templates_pivot, last_col)
    last_col += len(templates_pivot) * 3

    logging.info("process 'host' statistics data") 
    df_host = xl_create_dataframe(ws_host)
    host_pivot = ["ho_proxy", "ho_mode ajout", "ho_etat", "ho_maintenance"]
    for ti in host_title:
        if ti[0:8] in ["ma_{$TAG", "ma_{$SNM"]:
            host_pivot.append(ti)
        if ti in ["in_hardware", "in_type", "in_model", "in_os", "in_os_full", "in_software_app_a", 
            "in_location", "in_site_city", "in_contact"]:
            host_pivot.append(ti)
    host_value = "ho_host"    
    zbx_stat(df_host, ws_stat, host_value, host_pivot, last_col)
    last_col += len(host_pivot) * 3

    logging.info("end of statistics data process") 


    del wb["Sheet"]

    wb.save(dir_data + xlfile)


    # ending program

    logging.info("program ended")